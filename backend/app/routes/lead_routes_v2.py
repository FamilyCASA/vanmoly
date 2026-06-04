"""
留资引导模块路由 V2.0 - 多分店客资线索管理系统
API端点: /api/v3/leads
支持积分激励、公海机制、转化漏斗
"""
from flask import Blueprint, request, jsonify
import io
import os
from sqlalchemy import desc, asc, or_, func, and_
from datetime import datetime, timedelta, date
from app import db
from app.models.lead_v2 import Lead, LeadFollow, LeadPoint, LeadDistribution, LeadChannelStat
from app.routes.auth_routes_v2 import jwt_required_v2

lead_v2_bp = Blueprint('lead_v2', __name__, url_prefix='/api/v3')


def api_response(code=200, message='success', data=None):
    """统一API响应格式"""
    return jsonify({
        'code': code,
        'message': message,
        'data': data,
        'timestamp': int(datetime.utcnow().timestamp())
    }), code


# ========== 线索基础接口 ==========

@lead_v2_bp.route('/leads', methods=['POST'])
@jwt_required_v2
def create_lead(current_user, ):
    """
    创建线索（带积分）
    
    请求体:
    {
        "name": "张先生",
        "phone": "13800138000",
        "wechat": "wx123",
        "source": "案例留资",
        "building_name": "万科城",
        "building_address": "成都市高新区xxx",
        "house_type": "三室两厅",
        "area": 120,
        "budget": "20-30万",
        "intention_level": "高"
    }
    """
    try:
        data = request.get_json()
        if not data:
            return api_response(code=400, message='请求体不能为空')
        
        phone = data.get('phone', '').strip()
        if not phone or not phone.isdigit() or len(phone) != 11:
            return api_response(code=400, message='手机号格式不正确')
        
        # 检查手机号是否已存在
        existing = Lead.query.filter_by(phone=phone).first()
        if existing:
            return api_response(code=400, message='该手机号已存在')
        
        # 创建新线索
        current_user_id = request.current_user.get("id")
        lead = Lead(
            name=data.get('name', '').strip(),
            phone=phone,
            wechat=data.get('wechat'),
            gender=data.get('gender'),
            source=data.get('source', '其他'),
            source_detail=data.get('source_detail'),
            building_name=data.get('building_name'),
            building_address=data.get('building_address'),
            house_type=data.get('house_type'),
            area=data.get('area'),
            floor=data.get('floor'),
            delivery_date=datetime.strptime(data['delivery_date'], '%Y-%m-%d').date() if data.get('delivery_date') else None,
            decoration_status=data.get('decoration_status'),
            decoration_type=data.get('decoration_type'),
            style_preference=data.get('style_preference'),
            budget=data.get('budget'),
            timeline=data.get('timeline'),
            detailed_needs=data.get('detailed_needs'),
            family_structure=data.get('family_structure'),
            living_habits=data.get('living_habits'),
            hobbies=data.get('hobbies'),
            special_requirements=data.get('special_requirements'),
            focus_points=data.get('focus_points'),
            tags=data.get('tags', []),
            status=Lead.STATUS_PENDING,
            intention_level=data.get('intention_level', '中'),
            conversion_level=Lead.LEVEL_LEAD,
            created_by=current_user_id,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')[:500]
        )
        
        db.session.add(lead)
        db.session.flush()  # 获取lead.id
        
        # 添加积分：录入线索 +1分
        if current_user_id:
            LeadPoint.add_points(
                lead_id=lead.id,
                employee_id=current_user_id,
                point_type=LeadPoint.TYPE_CREATE,
                description=f'录入线索: {lead.name}'
            )
        
        db.session.commit()
        
        return api_response(
            code=201,
            message='线索创建成功，积分+1',
            data=lead.to_dict()
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'创建失败: {str(e)}')


@lead_v2_bp.route('/leads', methods=['GET'])
@jwt_required_v2
def get_leads(current_user, ):
    """
    获取线索列表（支持高级筛选）
    
    查询参数:
    - status: 转化状态筛选
    - intention_level: 意向等级筛选 (高/中/低)
    - conversion_level: 转化等级筛选 (线索/客户/VIP/SVIP)
    - assigned_to: 负责人ID
    - source: 来源渠道筛选
    - is_in_sea: 是否在公海 (true/false)
    - is_overdue: 是否逾期 (true/false)
    - keyword: 关键词搜索（姓名/手机号/楼盘）
    - building_name: 楼盘名称筛选
    - tags: 标签筛选（逗号分隔）
    - start_date: 开始日期
    - end_date: 结束日期
    - sort_by: 排序字段 (created_at/last_follow_at/next_follow_at/total_points)
    - sort_order: 排序方向 (asc/desc)
    - page: 页码，默认1
    - page_size: 每页数量，默认10
    """
    try:
        # 获取查询参数
        status = request.args.get('status')
        intention_level = request.args.get('intention_level')
        conversion_level = request.args.get('conversion_level')
        assigned_to = request.args.get('assigned_to', type=int)
        source = request.args.get('source')
        is_in_sea = request.args.get('is_in_sea')
        is_overdue = request.args.get('is_overdue')
        keyword = request.args.get('keyword', '').strip()
        building_name = request.args.get('building_name')
        tags = request.args.get('tags')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        sort_by = request.args.get('sort_by', 'created_at')
        sort_order = request.args.get('sort_order', 'desc')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 10, type=int)
        
        page_size = min(max(page_size, 1), 100)
        
        # 构建查询
        query = Lead.query
        
        # 状态筛选
        if status:
            query = query.filter(Lead.status == status)
        if intention_level:
            query = query.filter(Lead.intention_level == intention_level)
        if conversion_level:
            query = query.filter(Lead.conversion_level == conversion_level)
        if assigned_to:
            query = query.filter(Lead.assigned_to == assigned_to)
        if source:
            query = query.filter(Lead.source == source)
        if is_in_sea is not None:
            query = query.filter(Lead.is_in_sea == (is_in_sea.lower() == 'true'))
        if is_overdue is not None:
            query = query.filter(Lead.is_overdue == (is_overdue.lower() == 'true'))
        if building_name:
            query = query.filter(Lead.building_name.contains(building_name))
        if tags:
            tag_list = tags.split(',')
            for tag in tag_list:
                query = query.filter(Lead.tags.contains(tag))
        
        # 关键词搜索
        if keyword:
            query = query.filter(
                or_(
                    Lead.name.contains(keyword),
                    Lead.phone.contains(keyword),
                    Lead.building_name.contains(keyword)
                )
            )
        
        # 日期范围
        if start_date:
            query = query.filter(Lead.created_at >= f"{start_date} 00:00:00")
        if end_date:
            query = query.filter(Lead.created_at <= f"{end_date} 23:59:59")
        
        # 排序
        sort_column = getattr(Lead, sort_by, Lead.created_at)
        if sort_order == 'desc':
            query = query.order_by(desc(sort_column))
        else:
            query = query.order_by(asc(sort_column))
        
        # 分页
        pagination = query.paginate(page=page, per_page=page_size, error_out=False)
        
        return api_response(data={
            'items': [item.to_dict() for item in pagination.items],
            'total': pagination.total,
            'page': page,
            'page_size': page_size,
            'total_pages': pagination.pages
        })
        
    except Exception as e:
        return api_response(code=500, message=f'获取列表失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>', methods=['GET'])
@jwt_required_v2
def get_lead_detail(current_user, lead_id):
    """获取线索详情"""
    try:
        include_follows = request.args.get('include_follows', 'true').lower() == 'true'
        include_points = request.args.get('include_points', 'false').lower() == 'true'
        
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        return api_response(data=lead.to_dict(
            include_follows=include_follows,
            include_points=include_points
        ))
    except Exception as e:
        return api_response(code=500, message=f'获取详情失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>', methods=['PUT'])
@jwt_required_v2
def update_lead(current_user, lead_id):
    """更新线索信息"""
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        data = request.get_json()
        if not data:
            return api_response(code=400, message='请求体不能为空')
        
        # 可更新字段列表
        updatable_fields = [
            'name', 'wechat', 'gender', 'source', 'source_detail',
            'building_name', 'building_address', 'house_type', 'area', 'floor',
            'delivery_date', 'decoration_status', 'decoration_type',
            'style_preference', 'budget', 'timeline', 'detailed_needs',
            'family_structure', 'living_habits', 'hobbies',
            'special_requirements', 'focus_points', 'tags',
            'intention_level', 'remark'
        ]
        
        updated = False
        for field in updatable_fields:
            if field in data:
                if field == 'delivery_date' and data[field]:
                    setattr(lead, field, datetime.strptime(data[field], '%Y-%m-%d').date())
                else:
                    setattr(lead, field, data[field])
                updated = True
        
        if updated:
            lead.updated_at = datetime.utcnow()
            db.session.commit()
        
        return api_response(message='更新成功', data=lead.to_dict())
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'更新失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>', methods=['DELETE'])
@jwt_required_v2
def delete_lead(current_user, lead_id):
    """删除线索"""
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        db.session.delete(lead)
        db.session.commit()
        
        return api_response(message='删除成功')
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'删除失败: {str(e)}')


# ========== 快速跟进接口（带积分） ==========

@lead_v2_bp.route('/leads/<int:lead_id>/follow', methods=['POST'])
@jwt_required_v2
def add_follow(current_user, lead_id):
    """
    快速跟进（带积分）
    
    请求体:
    {
        "follow_type": "电话",
        "content": "客户意向强烈，约明天到店",
        "result": "意向强烈",
        "next_follow_at": "2026-04-27 10:00:00",
        "is_visited": true,
        "visited_at": "2026-04-27 14:00:00"
    }
    """
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        data = request.get_json()
        if not data or not data.get('content'):
            return api_response(code=400, message='跟进内容不能为空')
        
        current_user_id = request.current_user.get("id")
        
        # 解析时间
        next_follow_at = None
        if data.get('next_follow_at'):
            next_follow_at = datetime.strptime(data['next_follow_at'], '%Y-%m-%d %H:%M:%S')
        
        visited_at = None
        if data.get('visited_at'):
            visited_at = datetime.strptime(data['visited_at'], '%Y-%m-%d %H:%M:%S')
        
        # 添加跟进记录
        follow = LeadFollow(
            lead_id=lead_id,
            follow_type=data.get('follow_type', '其他'),
            content=data.get('content'),
            result=data.get('result'),
            next_follow_at=next_follow_at,
            is_visited=data.get('is_visited', False),
            visited_at=visited_at,
            attachments=data.get('attachments'),
            operator_id=current_user_id
        )
        db.session.add(follow)
        db.session.flush()
        
        # 更新线索跟进信息
        lead.follow_count += 1
        lead.last_follow_at = datetime.utcnow()
        lead.next_follow_at = next_follow_at
        lead.is_overdue = False
        lead.overdue_days = 0
        
        # 如果设置了预约到店
        if data.get('is_visited') and visited_at:
            lead.status = Lead.STATUS_VISITED
        elif lead.status == Lead.STATUS_PENDING:
            lead.status = Lead.STATUS_FOLLOWING
        
        db.session.flush()
        
        # 添加积分
        points_earned = 0
        point_records = []
        
        # 1. 有效跟进 +1分
        LeadPoint.add_points(
            lead_id=lead_id,
            employee_id=current_user_id,
            point_type=LeadPoint.TYPE_FOLLOW,
            description='有效跟进',
            follow_id=follow.id
        )
        points_earned += 1
        point_records.append({'type': '有效跟进', 'points': 1})
        
        # 2. 预约到店 +0.5分
        if data.get('is_visited'):
            LeadPoint.add_points(
                lead_id=lead_id,
                employee_id=current_user_id,
                point_type=LeadPoint.TYPE_APPOINT_VISIT,
                description='预约到店',
                follow_id=follow.id
            )
            points_earned += 0.5
            point_records.append({'type': '预约到店', 'points': 0.5})
        
        # 3. 获取真实需求 +1分（检查是否完善了需求信息）
        if lead.detailed_needs and lead.budget and lead.family_structure:
            # 检查是否已获得过该积分
            existing = LeadPoint.query.filter_by(
                lead_id=lead_id,
                employee_id=current_user_id,
                point_type=LeadPoint.TYPE_GET_NEEDS
            ).first()
            if not existing:
                LeadPoint.add_points(
                    lead_id=lead_id,
                    employee_id=current_user_id,
                    point_type=LeadPoint.TYPE_GET_NEEDS,
                    description='获取真实需求',
                    follow_id=follow.id
                )
                points_earned += 1
                point_records.append({'type': '获取需求', 'points': 1})
        
        db.session.commit()
        
        return api_response(
            code=201,
            message=f'跟进成功，获得{points_earned}积分',
            data={
                'follow': follow.to_dict(),
                'points_earned': points_earned,
                'point_records': point_records
            }
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'跟进失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>/visit', methods=['POST'])
@jwt_required_v2
def mark_visited(current_user, lead_id):
    """
    标记实际到店（+2分）
    
    请求体:
    {
        "visited_at": "2026-04-27 14:30:00"
    }
    """
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        data = request.get_json()
        visited_at = datetime.utcnow()
        if data and data.get('visited_at'):
            visited_at = datetime.strptime(data['visited_at'], '%Y-%m-%d %H:%M:%S')
        
        lead.is_visited = True
        lead.visited_at = visited_at
        lead.status = Lead.STATUS_VISITED
        db.session.flush()
        
        # 添加积分
        current_user_id = request.current_user.get("id")
        LeadPoint.add_points(
            lead_id=lead_id,
            employee_id=current_user_id,
            point_type=LeadPoint.TYPE_ACTUAL_VISIT,
            description='客户实际到店'
        )
        
        db.session.commit()
        
        return api_response(
            message='标记到店成功，积分+2',
            data=lead.to_dict()
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'标记失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>/deposit', methods=['POST'])
@jwt_required_v2
def mark_deposit(current_user, lead_id):
    """
    标记交定金（+10分）
    
    请求体:
    {
        "amount": 5000,
        "deposit_at": "2026-04-27"
    }
    """
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        data = request.get_json()
        
        lead.deposit_amount = data.get('amount', 0)
        if data.get('deposit_at'):
            lead.deposit_at = datetime.strptime(data['deposit_at'], '%Y-%m-%d')
        else:
            lead.deposit_at = datetime.utcnow()
        lead.status = Lead.STATUS_DEPOSIT
        lead.conversion_level = Lead.LEVEL_VIP
        db.session.flush()
        
        # 添加积分
        current_user_id = request.current_user.get("id")
        LeadPoint.add_points(
            lead_id=lead_id,
            employee_id=current_user_id,
            point_type=LeadPoint.TYPE_DEPOSIT,
            description=f'交定金 {lead.deposit_amount}元'
        )
        
        db.session.commit()
        
        return api_response(
            message='标记定金成功，积分+10',
            data=lead.to_dict()
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'标记失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>/contract', methods=['POST'])
@jwt_required_v2
def mark_contract(current_user, lead_id):
    """
    标记签约（根据类型加不同积分）
    
    请求体:
    {
        "contract_type": "签约全案",
        "amount": 300000,
        "contract_at": "2026-04-27"
    }
    """
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        data = request.get_json()
        contract_type = data.get('contract_type')
        
        if contract_type not in LeadPoint.POINT_RULES:
            return api_response(code=400, message=f'无效的签约类型')
        
        lead.contract_type = contract_type
        lead.contract_amount = data.get('amount', 0)
        if data.get('contract_at'):
            lead.contract_at = datetime.strptime(data['contract_at'], '%Y-%m-%d')
        else:
            lead.contract_at = datetime.utcnow()
        lead.status = Lead.STATUS_CONTRACTED
        lead.conversion_level = Lead.LEVEL_VIP
        db.session.flush()
        
        # 添加积分
        current_user_id = request.current_user.get("id")
        points = LeadPoint.POINT_RULES[contract_type]
        LeadPoint.add_points(
            lead_id=lead_id,
            employee_id=current_user_id,
            point_type=contract_type,
            description=f'{contract_type} {lead.contract_amount}元'
        )
        
        db.session.commit()
        
        return api_response(
            message=f'标记签约成功，积分+{points}',
            data=lead.to_dict()
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'标记失败: {str(e)}')


# ========== 公海接口 ==========

@lead_v2_bp.route('/leads/sea', methods=['GET'])
@jwt_required_v2
def get_sea_leads(current_user, ):
    """获取公海线索列表"""
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 10, type=int)
        keyword = request.args.get('keyword', '').strip()
        
        query = Lead.query.filter_by(is_in_sea=True)
        
        if keyword:
            query = query.filter(
                or_(
                    Lead.name.contains(keyword),
                    Lead.phone.contains(keyword),
                    Lead.building_name.contains(keyword)
                )
            )
        
        # 按入公海时间倒序
        query = query.order_by(desc(Lead.sea_at))
        
        pagination = query.paginate(page=page, per_page=page_size, error_out=False)
        
        return api_response(data={
            'items': [item.to_dict() for item in pagination.items],
            'total': pagination.total,
            'page': page,
            'page_size': page_size
        })
        
    except Exception as e:
        return api_response(code=500, message=f'获取公海列表失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>/retrieve', methods=['POST'])
@jwt_required_v2
def retrieve_from_sea(current_user, lead_id):
    """从公海领取线索"""
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        if not lead.is_in_sea:
            return api_response(code=400, message='该线索不在公海')
        
        current_user_id = request.current_user.get("id")
        
        # 检查每日领取上限（10条）
        today = date.today()
        retrieved_today = LeadDistribution.query.filter(
            and_(
                LeadDistribution.to_employee_id == current_user_id,
                LeadDistribution.distribution_type == '领取',
                func.date(LeadDistribution.created_at) == today
            )
        ).count()
        
        if retrieved_today >= 10:
            return api_response(code=400, message='今日领取已达上限（10条）')
        
        # 领取线索
        lead.retrieve_from_sea(current_user_id)
        
        # 记录分配
        distribution = LeadDistribution(
            lead_id=lead_id,
            from_employee_id=None,
            to_employee_id=current_user_id,
            distribution_type='领取',
            reason='从公海领取'
        )
        db.session.add(distribution)
        db.session.commit()
        
        return api_response(message='领取成功', data=lead.to_dict())
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'领取失败: {str(e)}')


@lead_v2_bp.route('/leads/<int:lead_id>/assign', methods=['POST'])
@jwt_required_v2
def assign_lead(current_user, lead_id):
    """
    分配线索给员工
    
    请求体:
    {
        "employee_id": 123,
        "reason": "手动分配"
    }
    """
    try:
        lead = Lead.query.get(lead_id)
        if not lead:
            return api_response(code=404, message='线索不存在')
        
        data = request.get_json()
        employee_id = data.get('employee_id')
        
        if not employee_id:
            return api_response(code=400, message='员工ID不能为空')
        
        current_user_id = request.current_user.get("id")
        old_assignee = lead.assigned_to
        
        # 分配线索
        lead.assign_to(employee_id, assigned_by=current_user_id)
        
        # 记录分配
        distribution = LeadDistribution(
            lead_id=lead_id,
            from_employee_id=old_assignee,
            to_employee_id=employee_id,
            distributed_by=current_user_id,
            distribution_type='手动分配',
            reason=data.get('reason', '手动分配')
        )
        db.session.add(distribution)
        db.session.commit()
        
        return api_response(message='分配成功', data=lead.to_dict())
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'分配失败: {str(e)}')


# ========== 批量操作接口 ==========

@lead_v2_bp.route('/leads/batch/assign', methods=['POST'])
@jwt_required_v2
def batch_assign(current_user, ):
    """批量分配线索"""
    try:
        data = request.get_json()
        lead_ids = data.get('lead_ids', [])
        employee_id = data.get('employee_id')
        
        if not lead_ids or not employee_id:
            return api_response(code=400, message='线索ID列表和员工ID不能为空')
        
        current_user_id = request.current_user.get("id")
        success_count = 0
        
        for lead_id in lead_ids:
            lead = Lead.query.get(lead_id)
            if lead:
                old_assignee = lead.assigned_to
                lead.assign_to(employee_id, assigned_by=current_user_id)
                
                distribution = LeadDistribution(
                    lead_id=lead_id,
                    from_employee_id=old_assignee,
                    to_employee_id=employee_id,
                    distributed_by=current_user_id,
                    distribution_type='批量分配',
                    reason='批量分配'
                )
                db.session.add(distribution)
                success_count += 1
        
        db.session.commit()
        
        return api_response(
            message=f'成功分配{success_count}条线索',
            data={'assigned_count': success_count}
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'批量分配失败: {str(e)}')


@lead_v2_bp.route('/leads/batch/invalid', methods=['POST'])
@jwt_required_v2
def batch_mark_invalid(current_user, ):
    """批量标记无效"""
    try:
        data = request.get_json()
        lead_ids = data.get('lead_ids', [])
        reason = data.get('reason', '')
        
        if not lead_ids:
            return api_response(code=400, message='线索ID列表不能为空')
        
        success_count = 0
        for lead_id in lead_ids:
            lead = Lead.query.get(lead_id)
            if lead:
                lead.is_invalid = True
                lead.invalid_reason = reason
                lead.status = Lead.STATUS_INVALID
                success_count += 1
        
        db.session.commit()
        
        return api_response(
            message=f'成功标记{success_count}条线索为无效',
            data={'invalid_count': success_count}
        )
        
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'批量标记失败: {str(e)}')


# ========== 积分排行榜接口 ==========

@lead_v2_bp.route('/points/ranking', methods=['GET'])
@jwt_required_v2
def get_points_ranking(current_user, ):
    """
    获取积分排行榜
    
    查询参数:
    - period: 周期 (today/week/month/all) 默认month
    - limit: 返回数量 默认10
    """
    try:
        period = request.args.get('period', 'month')
        limit = request.args.get('limit', 10, type=int)
        
        # 计算时间范围
        now = datetime.utcnow()
        if period == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'week':
            start_date = now - timedelta(days=7)
        elif period == 'month':
            start_date = now - timedelta(days=30)
        else:  # all
            start_date = datetime(2000, 1, 1)
        
        # 查询积分统计
        from app.models import Employee, Department
        
        ranking = db.session.query(
            Employee.id,
            Employee.name,
            Department.name.label('department_name'),
            func.coalesce(func.sum(LeadPoint.points), 0).label('total_points'),
            func.count(LeadPoint.id).label('point_count')
        ).outerjoin(
            LeadPoint, Employee.id == LeadPoint.employee_id
        ).outerjoin(
            Department, Employee.department_id == Department.id
        ).filter(
            LeadPoint.created_at >= start_date
        ).group_by(
            Employee.id
        ).order_by(
            desc('total_points')
        ).limit(limit).all()
        
        return api_response(data=[{
            'rank': i + 1,
            'employee_id': r.id,
            'employee_name': r.name,
            'department': r.department_name,
            'total_points': float(r.total_points),
            'point_count': r.point_count
        } for i, r in enumerate(ranking)])
        
    except Exception as e:
        return api_response(code=500, message=f'获取排行榜失败: {str(e)}')


@lead_v2_bp.route('/points/my', methods=['GET'])
@jwt_required_v2
def get_my_points(current_user, ):
    """获取我的积分明细"""
    try:
        current_user_id = request.current_user.get("id")
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)
        
        query = LeadPoint.query.filter_by(employee_id=current_user_id)
        query = query.order_by(desc(LeadPoint.created_at))
        
        pagination = query.paginate(page=page, per_page=page_size, error_out=False)
        
        # 统计今日、本周、本月、总积分
        now = datetime.utcnow()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=7)
        month_start = now - timedelta(days=30)
        
        today_points = db.session.query(func.coalesce(func.sum(LeadPoint.points), 0)).filter(
            LeadPoint.employee_id == current_user_id,
            LeadPoint.created_at >= today_start
        ).scalar()
        
        week_points = db.session.query(func.coalesce(func.sum(LeadPoint.points), 0)).filter(
            LeadPoint.employee_id == current_user_id,
            LeadPoint.created_at >= week_start
        ).scalar()
        
        month_points = db.session.query(func.coalesce(func.sum(LeadPoint.points), 0)).filter(
            LeadPoint.employee_id == current_user_id,
            LeadPoint.created_at >= month_start
        ).scalar()
        
        total_points = db.session.query(func.coalesce(func.sum(LeadPoint.points), 0)).filter(
            LeadPoint.employee_id == current_user_id
        ).scalar()
        
        return api_response(data={
            'stats': {
                'today': float(today_points),
                'week': float(week_points),
                'month': float(month_points),
                'total': float(total_points)
            },
            'items': [item.to_dict() for item in pagination.items],
            'total': pagination.total,
            'page': page,
            'page_size': page_size
        })
        
    except Exception as e:
        return api_response(code=500, message=f'获取积分明细失败: {str(e)}')


# ========== 统计接口 ==========

@lead_v2_bp.route('/stats/overview', methods=['GET'])
@jwt_required_v2
def get_stats_overview(current_user):
    """获取线索总览统计"""
    try:
        now = datetime.utcnow()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # 今日数据
        today_new = Lead.query.filter(Lead.created_at >= today_start).count()
        today_follow = LeadFollow.query.filter(LeadFollow.created_at >= today_start).count()
        today_visit = Lead.query.filter(Lead.visited_at >= today_start).count()
        today_deposit = Lead.query.filter(Lead.deposit_at >= today_start).count()
        today_contract = Lead.query.filter(Lead.contract_at >= today_start).count()
        
        # 总体数据
        total_leads = Lead.query.count()
        total_sea = Lead.query.filter_by(is_in_sea=True).count()
        total_overdue = Lead.query.filter_by(is_overdue=True).count()
        
        # 状态分布
        status_stats = db.session.query(
            Lead.status,
            func.count(Lead.id)
        ).group_by(Lead.status).all()
        
        # 等级分布
        level_stats = db.session.query(
            Lead.conversion_level,
            func.count(Lead.id)
        ).group_by(Lead.conversion_level).all()
        
        return api_response(data={
            'today': {
                'new': today_new,
                'follow': today_follow,
                'visit': today_visit,
                'deposit': today_deposit,
                'contract': today_contract
            },
            'overview': {
                'total_leads': total_leads,
                'total_sea': total_sea,
                'total_overdue': total_overdue
            },
            'status_distribution': {s: c for s, c in status_stats},
            'level_distribution': {l: c for l, c in level_stats}
        })
        
    except Exception as e:
        return api_response(code=500, message=f'获取统计失败: {str(e)}')


@lead_v2_bp.route('/stats/funnel', methods=['GET'])
@jwt_required_v2
def get_conversion_funnel(current_user, ):
    """获取转化漏斗"""
    try:
        # 各阶段数量
        total = Lead.query.count()
        following = Lead.query.filter(
            Lead.status.in_([Lead.STATUS_FOLLOWING, Lead.STATUS_VISITED, Lead.STATUS_MEASURED, 
                           Lead.STATUS_SCHEMED, Lead.STATUS_DEPOSIT, Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL])
        ).count()
        visited = Lead.query.filter(
            Lead.status.in_([Lead.STATUS_VISITED, Lead.STATUS_MEASURED, Lead.STATUS_SCHEMED, 
                           Lead.STATUS_DEPOSIT, Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL])
        ).count()
        measured = Lead.query.filter(
            Lead.status.in_([Lead.STATUS_MEASURED, Lead.STATUS_SCHEMED, Lead.STATUS_DEPOSIT, 
                           Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL])
        ).count()
        schemed = Lead.query.filter(
            Lead.status.in_([Lead.STATUS_SCHEMED, Lead.STATUS_DEPOSIT, Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL])
        ).count()
        deposit = Lead.query.filter(
            Lead.status.in_([Lead.STATUS_DEPOSIT, Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL])
        ).count()
        contracted = Lead.query.filter(
            Lead.status.in_([Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL])
        ).count()
        deal = Lead.query.filter_by(status=Lead.STATUS_DEAL).count()
        
        funnel = [
            {'stage': '线索', 'count': total, 'conversion_rate': 100},
            {'stage': '跟进', 'count': following, 'conversion_rate': round(following/total*100, 2) if total else 0},
            {'stage': '到店', 'count': visited, 'conversion_rate': round(visited/total*100, 2) if total else 0},
            {'stage': '量房', 'count': measured, 'conversion_rate': round(measured/total*100, 2) if total else 0},
            {'stage': '方案', 'count': schemed, 'conversion_rate': round(schemed/total*100, 2) if total else 0},
            {'stage': '定金', 'count': deposit, 'conversion_rate': round(deposit/total*100, 2) if total else 0},
            {'stage': '签约', 'count': contracted, 'conversion_rate': round(contracted/total*100, 2) if total else 0},
            {'stage': '成交', 'count': deal, 'conversion_rate': round(deal/total*100, 2) if total else 0}
        ]
        
        return api_response(data=funnel)
        
    except Exception as e:
        return api_response(code=500, message=f'获取漏斗失败: {str(e)}')


@lead_v2_bp.route('/stats/channels', methods=['GET'])
@jwt_required_v2
def get_channel_stats(current_user, ):
    """获取渠道统计"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = db.session.query(
            Lead.source,
            func.count(Lead.id).label('lead_count'),
            func.sum(func.case((Lead.is_visited == True, 1), else_=0)).label('visit_count'),
            func.sum(func.case((Lead.deposit_at != None, 1), else_=0)).label('deposit_count'),
            func.sum(func.case((Lead.contract_at != None, 1), else_=0)).label('contract_count'),
            func.coalesce(func.sum(Lead.contract_amount), 0).label('deal_amount')
        ).group_by(Lead.source)
        
        if start_date:
            query = query.filter(Lead.created_at >= f"{start_date} 00:00:00")
        if end_date:
            query = query.filter(Lead.created_at <= f"{end_date} 23:59:59")
        
        results = query.all()
        
        stats = []
        for r in results:
            conversion_rate = round(r.contract_count / r.lead_count * 100, 2) if r.lead_count else 0
            stats.append({
                'channel': r.source or '未知',
                'lead_count': r.lead_count,
                'visit_count': int(r.visit_count or 0),
                'deposit_count': int(r.deposit_count or 0),
                'contract_count': int(r.contract_count or 0),
                'deal_amount': float(r.deal_amount or 0),
                'conversion_rate': conversion_rate
            })
        
        # 按线索数排序
        stats.sort(key=lambda x: x['lead_count'], reverse=True)
        
        return api_response(data=stats)
        
    except Exception as e:
        return api_response(code=500, message=f'获取渠道统计失败: {str(e)}')


# ========== 筛选选项接口 ==========

@lead_v2_bp.route('/leads/filters', methods=['GET'])
@jwt_required_v2
def get_lead_filters(current_user):
    """获取线索筛选选项"""
    try:
        # 来源渠道
        sources = db.session.query(Lead.source).distinct().all()
        source_list = [s[0] for s in sources if s[0]]
        
        # 楼盘列表
        buildings = db.session.query(Lead.building_name).distinct().all()
        building_list = [b[0] for b in buildings if b[0]]
        
        # 标签
        all_tags = db.session.query(Lead.tags).filter(Lead.tags != None).all()
        tag_set = set()
        for tags in all_tags:
            if tags[0]:
                tag_set.update(tags[0])
        
        return api_response(data={
            'sources': source_list,
            'buildings': building_list,
            'tags': list(tag_set),
            'statuses': [
                Lead.STATUS_PENDING, Lead.STATUS_ASSIGNED, Lead.STATUS_FOLLOWING,
                Lead.STATUS_VISITED, Lead.STATUS_MEASURED, Lead.STATUS_SCHEMED,
                Lead.STATUS_DEPOSIT, Lead.STATUS_CONTRACTED, Lead.STATUS_DEAL,
                Lead.STATUS_INVALID, Lead.STATUS_SEA
            ],
            'intention_levels': [Lead.INTENTION_HIGH, Lead.INTENTION_MEDIUM, Lead.INTENTION_LOW],
            'conversion_levels': [Lead.LEVEL_LEAD, Lead.LEVEL_CUSTOMER, Lead.LEVEL_VIP, Lead.LEVEL_SVIP]
        })
        
    except Exception as e:
        return api_response(code=500, message=f'获取筛选选项失败: {str(e)}')


# ========== 待办提醒接口 ==========

@lead_v2_bp.route('/leads/todos', methods=['GET'])
@jwt_required_v2
def get_todos(current_user):
    """获取待办事项"""
    try:
        current_user_id = request.current_user.get("id")
        now = datetime.utcnow()
        today_end = now.replace(hour=23, minute=59, second=59)
        
        # 今日待跟进
        today_follow = Lead.query.filter(
            Lead.assigned_to == current_user_id,
            Lead.next_follow_at <= today_end,
            Lead.is_overdue == False
        ).count()
        
        # 逾期跟进
        overdue = Lead.query.filter(
            Lead.assigned_to == current_user_id,
            Lead.is_overdue == True
        ).count()
        
        # 待分配（店长/管理员）
        pending_assign = Lead.query.filter_by(status=Lead.STATUS_PENDING).count()
        
        # 待升级为客户
        pending_upgrade = Lead.query.filter(
            Lead.assigned_to == current_user_id,
            Lead.conversion_level == Lead.LEVEL_LEAD,
            Lead.building_name != None,
            Lead.building_address != None,
            Lead.detailed_needs != None
        ).count()
        
        return api_response(data={
            'today_follow': today_follow,
            'overdue': overdue,
            'pending_assign': pending_assign,
            'pending_upgrade': pending_upgrade
        })
        
    except Exception as e:
        return api_response(code=500, message=f'获取待办失败: {str(e)}')


# ========== Excel 批量导入线索 ==========

@lead_v2_bp.route('/leads/import-template', methods=['GET'])
@jwt_required_v2
def download_lead_import_template(current_user):
    """下载线索导入 Excel 模板"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '线索导入模板'
    
    # 表头样式
    header_fill = PatternFill(start_color='8B5A2B', end_color='8B5A2B', fill_type='solid')
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    cell_font = Font(name='微软雅黑', size=10)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    headers = ['姓名*', '手机号*', '微信号', '意向楼盘', '户型', '面积(㎡)', '预算', '风格偏好', '来源渠道', '意向等级', '装修状态', '备注']
    col_widths = [12, 15, 15, 18, 12, 10, 12, 12, 12, 10, 12, 20]
    
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    
    ws.row_dimensions[1].height = 28
    
    # 示例行
    example_data = ['张三', '13800138000', 'wx123456', '万科城', '三室两厅', '120', '25-30万', '现代简约', '案例留资', '高', '毛坯', '客户刚需，急装']
    for col, val in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.freeze_panes = 'A2'
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='线索导入模板.xlsx'
    )


@lead_v2_bp.route('/leads/import-excel', methods=['POST'])
@jwt_required_v2
def import_leads_from_excel(current_user):
    """Excel 批量导入线索
    
    Excel 列头支持（自动识别）：
    姓名/客户姓名, 手机号/电话/手机, 微信号, 
    意向楼盘/楼盘, 户型, 面积, 预算,
    风格偏好, 来源渠道/来源, 意向等级/等级(高/中/低),
    装修状态, 备注/备注信息
    """
    import io
    import openpyxl
    
    if 'file' not in request.files:
        return api_response(code=400, message='请上传 Excel 文件')
    
    file = request.files['file']
    if not file.filename:
        return api_response(code=400, message='文件名为空')
    
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {'.xlsx', '.xls'}:
        return api_response(code=400, message=f'不支持的文件格式 {ext}，请上传 xlsx/xls')
    
    try:
        # 读取 Excel
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows or len(rows) < 2:
            return api_response(code=400, message='Excel 文件为空或无数据')
        
        # 解析表头
        headers = [str(c).strip() if c else '' for c in rows[0]]
        
        # 列名映射
        COLUMN_MAP = {
            '姓名': 'name', '客户姓名': 'name', '客户': 'name',
            '手机号': 'phone', '电话': 'phone', '手机': 'phone', '联系电话': 'phone',
            '微信号': 'wechat',
            '意向楼盘': 'building_name', '楼盘': 'building_name', '楼盘名称': 'building_name',
            '户型': 'house_type',
            '面积': 'area', '面积(㎡)': 'area',
            '预算': 'budget', '装修预算': 'budget',
            '风格偏好': 'style_preference', '风格': 'style_preference',
            '来源渠道': 'source', '来源': 'source', '渠道': 'source',
            '意向等级': 'intention_level', '等级': 'intention_level', '意向': 'intention_level',
            '装修状态': 'decoration_status',
            '备注': 'remark', '备注信息': 'remark',
        }
        
        col_map = {}
        for i, h in enumerate(headers):
            h = str(h).strip()
            if h in COLUMN_MAP:
                col_map[i] = COLUMN_MAP[h]
        
        if 'name' not in col_map.values() and 'phone' not in col_map.values():
            return api_response(code=400, message='Excel 表头无法识别，请使用标准模板列名')
        
        imported = 0
        skipped = 0
        errors = []
        
        current_user_id = request.current_user.get('id')
        
        for idx, row in enumerate(rows[1:], 2):  # 从第2行开始
            try:
                row_data = {}
                for i in col_map:
                    val = row[i] if i < len(row) else None
                    row_data[col_map[i]] = str(val).strip() if val is not None else ''
                
                name = row_data.get('name', '').strip()
                phone = row_data.get('phone', '').strip()
                
                if not phone:
                    skipped += 1
                    continue
                
                if not phone.isdigit() or len(phone) < 11:
                    errors.append(f'第{idx}行：手机号格式错误')
                    skipped += 1
                    continue
                
                # 检查是否已存在
                existing = Lead.query.filter_by(phone=phone).first()
                if existing:
                    errors.append(f'第{idx}行：手机号 {phone} 已存在')
                    skipped += 1
                    continue
                
                # 生成编号
                today = datetime.now().strftime('%Y%m%d')
                seq = db.session.query(func.count(Lead.id)).scalar() or 0
                lead_no = f'LD{today}{seq + imported + 1:04d}'
                
                # 处理意向等级
                intention = row_data.get('intention_level', '中').strip()
                if intention not in ('高', '中', '低'):
                    intention = '中'
                
                # 处理来源
                source = row_data.get('source', '其他').strip() or '其他'
                
                # 处理面积
                area = None
                area_str = row_data.get('area', '').strip()
                if area_str and area_str.replace('.', '').isdigit():
                    area = float(area_str)
                
                lead = Lead(
                    lead_no=lead_no,
                    name=name or '未命名',
                    phone=phone,
                    wechat=row_data.get('wechat') or None,
                    building_name=row_data.get('building_name') or None,
                    house_type=row_data.get('house_type') or None,
                    area=area,
                    budget=row_data.get('budget') or None,
                    style_preference=row_data.get('style_preference') or None,
                    source=source,
                    intention_level=intention,
                    decoration_status=row_data.get('decoration_status') or None,
                    remark=row_data.get('remark') or None,
                    status=Lead.STATUS_PENDING,
                    conversion_level=Lead.LEVEL_LEAD,
                    created_by=current_user_id,
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent', '')[:500]
                )
                
                db.session.add(lead)
                imported += 1
                
            except Exception as e:
                errors.append(f'第{idx}行：{str(e)}')
                skipped += 1
        
        db.session.commit()
        
        msg = f'导入完成：成功 {imported} 条'
        if skipped > 0:
            msg += f'，跳过 {skipped} 条'
        if errors:
            msg += f'，错误：' + '；'.join(errors[:5])
            if len(errors) > 5:
                msg += f'... 共{len(errors)}条错误'
        
        return api_response(code=200, message=msg, data={'imported': imported, 'skipped': skipped})
    
    except Exception as e:
        db.session.rollback()
        return api_response(code=500, message=f'导入失败: {str(e)}')
