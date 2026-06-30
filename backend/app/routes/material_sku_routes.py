"""
物料管理模块 - API路由
V3.0 全新设计
"""
import io
import os
from flask import Blueprint, request, jsonify
from app import db
from app.models.material_sku import (
    MaterialSKU, MaterialCategory, MaterialVariant, MaterialSupplier,
    CALC_TYPES, CUSTOMIZATION_RULE_TYPES, SKU_STATUS
)
from app.routes.auth_routes_v2 import jwt_required_v2
from datetime import datetime

material_sku_bp = Blueprint('material_sku', __name__, url_prefix='/api/v3/materials')


# ========== 分类管理 ==========

@material_sku_bp.route('/categories', methods=['GET'])
def get_categories():
    """获取分类树（公开接口）"""
    from sqlalchemy import or_ as sql_or
    # 支持 is_deleted=False 和 is_deleted=NULL 两种情况（迁移数据可能为NULL）
    categories = MaterialCategory.query.filter(
        MaterialCategory.parent_id == None,
        sql_or(MaterialCategory.is_deleted == False, MaterialCategory.is_deleted.is_(None)),
        MaterialCategory.is_enabled == True
    ).order_by(MaterialCategory.sort_order).all()
    return jsonify({
        'code': 200,
        'data': [c.to_dict() for c in categories]
    })



@material_sku_bp.route('/categories-with-materials', methods=['GET'])
def get_categories_with_materials():
    """获取有实际物料的分类树（用于物料选择下拉）"""
    # 查询所有有物料的L2分类ID
    from sqlalchemy import func, distinct
    
    l2_ids_with_materials = db.session.query(distinct(MaterialSKU.category_id)).filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.category_id.isnot(None)
    ).all()
    l2_id_set = {r[0] for r in l2_ids_with_materials}
    
    # 查询这些L2分类及其L1父分类
    l2_cats = MaterialCategory.query.filter(
        MaterialCategory.id.in_(l2_id_set),
        MaterialCategory.is_deleted == False
    ).all() if l2_id_set else []
    
    # 收集所有L1 ID
    l1_ids = {c.parent_id for c in l2_cats if c.parent_id}
    
    # 查询L1分类
    l1_cats = MaterialCategory.query.filter(
        MaterialCategory.id.in_(l1_ids),
        MaterialCategory.is_deleted == False
    ).order_by(MaterialCategory.sort_order).all() if l1_ids else []
    
    # 构建树
    result = []
    for l1 in l1_cats:
        children = [c.to_dict() for c in sorted(
            [c for c in l2_cats if c.parent_id == l1.id],
            key=lambda x: x.sort_order
        )]
        l1_dict = l1.to_dict()
        l1_dict['children'] = children
        result.append(l1_dict)
    
    # 也要包含只有L1级别物料的分类（没有L2子分类但L1本身有物料）
    l1_only_ids = db.session.query(distinct(MaterialSKU.category_id)).join(
        MaterialCategory, MaterialSKU.category_id == MaterialCategory.id
    ).filter(
        MaterialSKU.is_deleted == False,
        MaterialCategory.level == 1,
        MaterialCategory.is_deleted == False
    ).all()
    
    existing_l1_ids = {l1.id for l1 in l1_cats}
    for row in l1_only_ids:
        if row[0] not in existing_l1_ids:
            cat = MaterialCategory.query.get(row[0])
            if cat and not cat.is_deleted:
                cat_dict = cat.to_dict()
                cat_dict['children'] = []
                result.append(cat_dict)
    
    return jsonify({
        'code': 200,
        'data': result
    })


@material_sku_bp.route('/material-categories/processes', methods=['GET'])
@jwt_required_v2
def get_process_categories(current_user):
    """获取工艺增项分类列表
    特殊工艺是挂在每个一级分类下的二级分类，按 parent_id（一级分类ID）过滤
    不传 parent_id 时返回所有特殊工艺子分类（向后兼容）
    """
    from sqlalchemy import or_ as sql_or
    parent_id = request.args.get('parent_id', type=int)

    # 找到所有 name='特殊工艺' 的二级分类（parent_id != None）
    query = MaterialCategory.query.filter(
        MaterialCategory.name == '特殊工艺',
        MaterialCategory.parent_id != None,
        sql_or(MaterialCategory.is_deleted == False, MaterialCategory.is_deleted.is_(None)),
        MaterialCategory.is_enabled == True
    )
    if parent_id:
        query = query.filter(MaterialCategory.parent_id == parent_id)
    special_cats = query.order_by(MaterialCategory.sort_order).all()

    result = []
    for sp in special_cats:
        children = MaterialCategory.query.filter(
            MaterialCategory.parent_id == sp.id,
            sql_or(MaterialCategory.is_deleted == False, MaterialCategory.is_deleted.is_(None)),
            MaterialCategory.is_enabled == True
        ).order_by(MaterialCategory.sort_order).all()
        for child in children:
            remark = child.remark or ''
            coef = 1.0
            unit = ''
            unit_price = 0.0
            try:
                import re
                m = re.search(r'系数[:：]?\s*([\d.]+)', remark)
                if m:
                    coef = float(m.group(1))
                m = re.search(r'单位[:：]?\s*(\S+)', remark)
                if m:
                    unit = m.group(1)
                m = re.search(r'单价[:：]?\s*([\d.]+)', remark)
                if m:
                    unit_price = float(m.group(1))
            except Exception:
                pass
            result.append({
                'id': child.id,
                'name': child.name,
                'parent_id': sp.parent_id,   # 一级分类ID，前端可用于校验
                'coefficient': coef,
                'unit': unit,
                'unit_price': unit_price,
                'remark': child.remark,
            })

    return jsonify({'code': 200, 'data': result})


@material_sku_bp.route('/categories', methods=['POST'])
@jwt_required_v2
def create_category(current_user):
    """创建分类"""
    data = request.get_json()

    category = MaterialCategory(
        name=data['name'],
        code=data.get('code'),
        parent_id=data.get('parent_id'),
        level=data.get('level', 1),
        icon=data.get('icon'),
        color=data.get('color', '#8B5A2B'),
        sort_order=data.get('sort_order', 0)
    )

    db.session.add(category)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '创建成功',
        'data': category.to_dict()
    })


@material_sku_bp.route('/categories/<int:id>', methods=['PUT'])
@jwt_required_v2
def update_category(current_user, id):
    """更新分类"""
    category = MaterialCategory.query.get_or_404(id)
    data = request.get_json()

    category.name = data.get('name', category.name)
    category.code = data.get('code', category.code)
    category.icon = data.get('icon', category.icon)
    category.color = data.get('color', category.color)
    category.sort_order = data.get('sort_order', category.sort_order)
    category.is_enabled = data.get('is_enabled', category.is_enabled)

    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '更新成功',
        'data': category.to_dict()
    })


@material_sku_bp.route('/categories/<int:id>', methods=['DELETE'])
@jwt_required_v2
def delete_category(current_user, id):
    """删除分类"""
    category = MaterialCategory.query.get_or_404(id)

    # 检查是否有子分类
    if category.children:
        return jsonify({'code': 400, 'message': '请先删除子分类'}), 400

    # 检查是否有物料
    if category.materials.count() > 0:
        return jsonify({'code': 400, 'message': '该分类下还有物料，无法删除'}), 400

    category.is_deleted = True
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


# ========== 物料管理 ==========

@material_sku_bp.route('', methods=['GET'])
def get_materials():
    """获取物料列表（公开接口）"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '').strip()
    category_id = request.args.get('category_id', '')
    # 支持逗号分隔的多个 category_id（一级分类搜索其下所有二级分类）
    category_ids_raw = request.args.get('category_ids', '')
    status = request.args.get('status', '').strip()
    brand = request.args.get('brand', '').strip()
    unit = request.args.get('unit', '').strip()
    supply_chain = request.args.get('supply_chain', '').strip()
    env_level = request.args.get('env_level', '').strip()
    is_public = request.args.get('is_public', '', type=str)
    low_stock = request.args.get('low_stock', type=bool)

    query = MaterialSKU.query.filter_by(is_deleted=False)

    # 前台只展示公开物料
    if is_public in ('1', 'true'):
        query = query.filter_by(is_public=True)

    if status:
        query = query.filter_by(status=status)

    if brand:
        query = query.filter(MaterialSKU.brand.contains(brand))

    if unit:
        query = query.filter_by(unit=unit)

    if supply_chain:
        query = query.filter(MaterialSKU.supply_chain.contains(supply_chain))

    if env_level:
        query = query.filter_by(env_level=env_level)

    if keyword:
        query = query.filter(
            db.or_(
                MaterialSKU.name.contains(keyword),
                MaterialSKU.sku_code.contains(keyword),
                MaterialSKU.brand.contains(keyword)
            )
        )

    if category_id:
        try:
            cid = int(category_id)
            query = query.filter_by(category_id=cid)
        except ValueError:
            pass

    if category_ids_raw:
        # 支持逗号分隔的多个ID，用于一级分类搜索其下所有二级分类的物料
        try:
            cids = [int(x.strip()) for x in category_ids_raw.split(',') if x.strip()]
            if cids:
                query = query.filter(MaterialSKU.category_id.in_(cids))
        except ValueError:
            pass

    if low_stock:
        query = query.filter(MaterialSKU.stock_quantity <= MaterialSKU.stock_warning)

    query = query.order_by(MaterialSKU.created_at.desc())
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)

    return jsonify({
        'code': 200,
        'data': {
            'items': [m.to_dict() for m in pagination.items],
            'total': pagination.total,
            'page': page,
            'page_size': page_size
        }
    })


@material_sku_bp.route('/<int:id>', methods=['GET'])
def get_material(id):
    """获取物料详情（公开接口）"""
    material = MaterialSKU.query.get_or_404(id)
    return jsonify({
        'code': 200,
        'data': material.to_dict(include_variants=True)
    })


@material_sku_bp.route('', methods=['POST'])
@jwt_required_v2
def create_material(current_user):
    """创建物料"""
    data = request.get_json()

    # 生成SKU编码
    sku_code = data.get('sku_code') or generate_sku_code()

    material = MaterialSKU(
        sku_code=sku_code,
        name=data['name'],
        category_id=data.get('category_id'),
        brand=data.get('brand'),
        model=data.get('model'),
        specification=data.get('specification'),
        material=data.get('material'),
        origin=data.get('origin'),
        main_image=data.get('main_image'),
        images=data.get('images', []),
        cost_price=data.get('cost_price', 0),
        sale_price=data.get('sale_price', 0),
        market_price=data.get('market_price'),
        unit=data.get('unit', '件'),
        calc_type=data.get('calc_type', 'quantity'),
        stock_quantity=data.get('stock_quantity', 0),
        stock_warning=data.get('stock_warning', 10),
        customization_rules=data.get('customization_rules', []),
        has_variants=data.get('has_variants', False),
        variant_options=data.get('variant_options', []),
        has_craft_parts=data.get('has_craft_parts', False),
        craft_parts=data.get('craft_parts', []),
        description=data.get('description'),
        tags=data.get('tags', []),
        status=data.get('status', 'active'),
        color_name=data.get('color_name', ''),
        env_level=data.get('env_level', '合格'),
        supply_chain=data.get('supply_chain', '直供'),
        supplier_id=data.get('supplier_id') or 2,  # 默认供应商: 帝标家居
        created_by=1
    )
    # 品牌默认值
    if not material.brand:
        material.brand = 'D&B'

    db.session.add(material)
    db.session.flush()

    # 创建变体
    if data.get('variants'):
        for v_data in data['variants']:
            variant = MaterialVariant(
                sku_id=material.id,
                variant_code=v_data.get('variant_code'),
                variant_name=v_data.get('variant_name'),
                variant_values=v_data.get('variant_values'),
                image=v_data.get('image'),
                price_adjustment=v_data.get('price_adjustment', 0),
                stock_quantity=v_data.get('stock_quantity', 0)
            )
            db.session.add(variant)

    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '创建成功',
        'data': material.to_dict(include_variants=True)
    })


@material_sku_bp.route('/<int:id>', methods=['PUT'])
@jwt_required_v2
def update_material(current_user, id):
    """更新物料"""
    material = MaterialSKU.query.get_or_404(id)
    data = request.get_json()

    # 更新基础字段
    fields = [
        'name', 'category_id', 'brand', 'model', 'specification',
        'material', 'origin', 'main_image', 'images', 'cost_price',
        'sale_price', 'market_price', 'unit', 'calc_type',
        'stock_quantity', 'stock_warning', 'customization_rules',
        'has_variants', 'variant_options', 'has_craft_parts',
        'craft_parts', 'description', 'tags', 'status',
        'color_name', 'env_level', 'supply_chain', 'supplier_id'
    ]

    for field in fields:
        if field in data:
            setattr(material, field, data[field])

    # 更新变体
    if 'variants' in data:
        # 删除旧变体
        MaterialVariant.query.filter_by(sku_id=id).delete()
        # 创建新变体
        for v_data in data['variants']:
            variant = MaterialVariant(
                sku_id=material.id,
                variant_code=v_data.get('variant_code'),
                variant_name=v_data.get('variant_name'),
                variant_values=v_data.get('variant_values'),
                image=v_data.get('image'),
                price_adjustment=v_data.get('price_adjustment', 0),
                stock_quantity=v_data.get('stock_quantity', 0)
            )
            db.session.add(variant)

    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '更新成功',
        'data': material.to_dict(include_variants=True)
    })


@material_sku_bp.route('/<int:id>', methods=['DELETE'])
@jwt_required_v2
def delete_material(current_user, id):
    """删除物料"""
    material = MaterialSKU.query.get_or_404(id)
    material.is_deleted = True
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


# ========== 变体管理 ==========

@material_sku_bp.route('/<int:sku_id>/variants', methods=['POST'])
@jwt_required_v2
def add_variant(current_user, sku_id):
    """添加变体"""
    data = request.get_json()

    variant = MaterialVariant(
        sku_id=sku_id,
        variant_code=data.get('variant_code'),
        variant_name=data.get('variant_name'),
        variant_values=data.get('variant_values'),
        image=data.get('image'),
        price_adjustment=data.get('price_adjustment', 0),
        stock_quantity=data.get('stock_quantity', 0)
    )

    db.session.add(variant)

    # 更新SKU的变体状态
    sku = MaterialSKU.query.get(sku_id)
    sku.has_variants = True

    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '添加成功',
        'data': variant.to_dict()
    })


@material_sku_bp.route('/variants/<int:id>', methods=['PUT'])
@jwt_required_v2
def update_variant(current_user, id):
    """更新变体"""
    variant = MaterialVariant.query.get_or_404(id)
    data = request.get_json()

    variant.variant_code = data.get('variant_code', variant.variant_code)
    variant.variant_name = data.get('variant_name', variant.variant_name)
    variant.variant_values = data.get('variant_values', variant.variant_values)
    variant.image = data.get('image', variant.image)
    variant.price_adjustment = data.get('price_adjustment', variant.price_adjustment)
    variant.stock_quantity = data.get('stock_quantity', variant.stock_quantity)
    variant.is_enabled = data.get('is_enabled', variant.is_enabled)

    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '更新成功',
        'data': variant.to_dict()
    })


@material_sku_bp.route('/variants/<int:id>', methods=['DELETE'])
@jwt_required_v2
def delete_variant(current_user, id):
    """删除变体"""
    variant = MaterialVariant.query.get_or_404(id)
    db.session.delete(variant)
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


# ========== 供应商管理 ==========

@material_sku_bp.route('/suppliers', methods=['GET'])
@jwt_required_v2
def get_suppliers(current_user):
    """获取供应商列表（支持分页）"""
    keyword = request.args.get('keyword', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    
    query = MaterialSupplier.query.filter_by(is_deleted=False)
    
    if keyword:
        query = query.filter(
            db.or_(
                MaterialSupplier.name.contains(keyword),
                MaterialSupplier.contact_person.contains(keyword),
                MaterialSupplier.phone.contains(keyword),
                MaterialSupplier.brand.contains(keyword),
                MaterialSupplier.main_products.contains(keyword)
            )
        )
    
    if status:
        query = query.filter_by(status=status)
    
    total = query.count()
    suppliers = query.order_by(MaterialSupplier.created_at.desc())         .offset((page - 1) * page_size).limit(page_size).all()
    
    # 获取员工名称映射
    from app.models.auth_v2 import UserV2
    specialist_ids = [s.specialist_id for s in suppliers if s.specialist_id]
    specialist_map = {}
    if specialist_ids:
        users = UserV2.query.filter(UserV2.id.in_(specialist_ids)).all()
        specialist_map = {u.id: u.display_name or u.username for u in users}
    
    # 物料数量映射（通过 supplier_id 关联）
    from sqlalchemy import func
    count_rows = db.session.query(
        MaterialSKU.supplier_id,
        func.count(MaterialSKU.id)
    ).filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.supplier_id.isnot(None)
    ).group_by(MaterialSKU.supplier_id).all()
    count_map = {r[0]: r[1] for r in count_rows}
    
    data = []
    for s in suppliers:
        item = s.to_dict()
        item['specialist_name'] = specialist_map.get(s.specialist_id, '')
        item['material_count'] = count_map.get(s.id, 0)
        data.append(item)
    
    return jsonify({
        'code': 200,
        'data': {
            'items': data,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@material_sku_bp.route('/suppliers/<int:id>/materials', methods=['GET'])
@jwt_required_v2
def get_supplier_materials(current_user, id):
    """获取供应商关联物料列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '').strip()
    
    query = MaterialSKU.query.filter_by(is_deleted=False, supplier_id=id)
    if keyword:
        query = query.filter(MaterialSKU.name.contains(keyword))
    
    total = query.count()
    items = query.order_by(MaterialSKU.created_at.desc())         .offset((page - 1) * page_size).limit(page_size).all()
    
    return jsonify({
        'code': 200,
        'data': {
            'items': [item.to_dict() for item in items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@material_sku_bp.route('/suppliers/<int:id>', methods=['PUT'])
@jwt_required_v2
def update_supplier(current_user, id):
    """更新供应商（供应链登记）"""
    supplier = MaterialSupplier.query.get_or_404(id)
    data = request.get_json()
    
    from datetime import datetime as dt
    if 'cooperation_date' in data:
        if data['cooperation_date']:
            try:
                supplier.cooperation_date = dt.strptime(data['cooperation_date'], '%Y-%m-%d').date()
            except ValueError:
                supplier.cooperation_date = None
        else:
            supplier.cooperation_date = None
    
    supplier.name = data.get('name', supplier.name)
    supplier.brand = data.get('brand', supplier.brand)
    supplier.main_products = data.get('main_products', supplier.main_products)
    supplier.factory_address = data.get('factory_address', supplier.factory_address)
    supplier.store_address = data.get('store_address', supplier.store_address)
    supplier.contact_person = data.get('contact_person', supplier.contact_person)
    supplier.phone = data.get('phone', supplier.phone)
    supplier.email = data.get('email', supplier.email)
    supplier.specialist_id = data.get('specialist_id', supplier.specialist_id)
    supplier.status = data.get('status', supplier.status)
    supplier.level = data.get('level', supplier.level)
    supplier.payment_method = data.get('payment_method', supplier.payment_method)
    supplier.bank_account = data.get('bank_account', supplier.bank_account)
    supplier.bank_name = data.get('bank_name', supplier.bank_name)
    supplier.tax_number = data.get('tax_number', supplier.tax_number)
    supplier.address = data.get('address', supplier.address)
    supplier.remark = data.get('remark', supplier.remark)
    
    db.session.commit()
    return jsonify({
        'code': 200,
        'message': '更新成功',
        'data': supplier.to_dict()
    })


@material_sku_bp.route('/suppliers/<int:id>', methods=['DELETE'])
@jwt_required_v2
def delete_supplier(current_user, id):
    """删除供应商"""
    supplier = MaterialSupplier.query.get_or_404(id)
    supplier.is_deleted = True
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@material_sku_bp.route('/suppliers', methods=['POST'])
@jwt_required_v2
def create_supplier(current_user):
    """创建供应商（供应链登记）"""
    data = request.get_json()

    # 自动生成供应商编号
    count = MaterialSupplier.query.count()
    supplier_code = f'SUP-{str(count + 1).zfill(4)}'
    
    from datetime import datetime as dt
    cooperation_date = None
    if data.get('cooperation_date'):
        try:
            cooperation_date = dt.strptime(data['cooperation_date'], '%Y-%m-%d').date()
        except ValueError:
            pass

    supplier = MaterialSupplier(
        supplier_code=supplier_code,
        name=data['name'],
        brand=data.get('brand', ''),
        main_products=data.get('main_products', ''),
        factory_address=data.get('factory_address', ''),
        store_address=data.get('store_address', ''),
        contact_person=data.get('contact_person', ''),
        phone=data.get('phone', ''),
        email=data.get('email', ''),
        specialist_id=data.get('specialist_id'),
        status=data.get('status', 'active'),
        level=data.get('level', 'B'),
        cooperation_date=cooperation_date,
        payment_method=data.get('payment_method', ''),
        bank_account=data.get('bank_account', ''),
        bank_name=data.get('bank_name', ''),
        tax_number=data.get('tax_number', ''),
        address=data.get('address', ''),
        remark=data.get('remark', '')
    )

    db.session.add(supplier)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '创建成功',
        'data': supplier.to_dict()
    })


# ========== 选项字典 ==========

@material_sku_bp.route('/options', methods=['GET'])
@jwt_required_v2
def get_options(current_user):
    """获取选项字典"""
    return jsonify({
        'code': 200,
        'data': {
            'calc_types': CALC_TYPES,
            'customization_rule_types': CUSTOMIZATION_RULE_TYPES,
            'sku_status': SKU_STATUS
        }
    })


# ========== 分类统计 ==========

@material_sku_bp.route('/category-stats', methods=['GET'])
@jwt_required_v2
def get_category_stats(current_user):
    """获取分类物料统计"""
    # 获取所有分类
    categories = MaterialCategory.query.filter_by(is_deleted=False).all()
    
    stats = {}
    for cat in categories:
        # 获取该分类及其子分类的所有物料
        cat_ids = [cat.id] + [c.id for c in cat.children if not c.is_deleted]
        
        total = MaterialSKU.query.filter(
            MaterialSKU.category_id.in_(cat_ids),
            MaterialSKU.is_deleted == False
        ).count()
        
        active = MaterialSKU.query.filter(
            MaterialSKU.category_id.in_(cat_ids),
            MaterialSKU.is_deleted == False,
            MaterialSKU.status == 'active'
        ).count()
        
        low_stock = MaterialSKU.query.filter(
            MaterialSKU.category_id.in_(cat_ids),
            MaterialSKU.is_deleted == False,
            MaterialSKU.stock_quantity <= MaterialSKU.stock_warning
        ).count()
        
        stats[cat.id] = {
            'total': total,
            'active': active,
            'low_stock': low_stock
        }
    
    return jsonify({'code': 200, 'data': stats})


# ========== 供应商统计 ==========

@material_sku_bp.route('/supplier-stats', methods=['GET'])
@jwt_required_v2
def get_supplier_stats(current_user):
    """获取供应商统计"""
    total = MaterialSupplier.query.filter_by(is_deleted=False).count()
    
    status_stats = db.session.query(
        MaterialSupplier.status,
        db.func.count(MaterialSupplier.id)
    ).filter_by(is_deleted=False).group_by(MaterialSupplier.status).all()
    
    return jsonify({
        'code': 200,
        'data': {
            'total': total,
            'active': dict(status_stats).get('active', 0),
            'paused': dict(status_stats).get('paused', 0),
            'terminated': dict(status_stats).get('terminated', 0),
            'material_count': MaterialSKU.query.filter_by(is_deleted=False).count()
        }
    })


@material_sku_bp.route('/suppliers/options', methods=['GET'])
@jwt_required_v2
def get_supplier_options(current_user):
    """获取供应商下拉选项列表（精简版）"""
    suppliers = MaterialSupplier.query.filter_by(is_deleted=False).order_by(MaterialSupplier.name).all()
    return jsonify({
        'code': 200,
        'data': [{'id': s.id, 'name': s.name, 'brand': s.brand, 'supplier_code': s.supplier_code} for s in suppliers]
    })


# ========== 统计 ==========

@material_sku_bp.route('/stats', methods=['GET'])
@jwt_required_v2
def get_stats(current_user):
    """获取物料统计"""
    # 总物料数
    total = MaterialSKU.query.filter_by(is_deleted=False).count()

    # 按状态统计
    status_stats = db.session.query(
        MaterialSKU.status,
        db.func.count(MaterialSKU.id)
    ).filter_by(is_deleted=False).group_by(MaterialSKU.status).all()

    # 低库存物料
    low_stock = MaterialSKU.query.filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.stock_quantity <= MaterialSKU.stock_warning
    ).count()

    # 按分类统计
    category_stats = db.session.query(
        MaterialCategory.name,
        db.func.count(MaterialSKU.id)
    ).join(MaterialSKU, MaterialCategory.id == MaterialSKU.category_id).filter(
        MaterialSKU.is_deleted == False
    ).group_by(MaterialCategory.name).all()

    # 计算今日新增（过去24小时）
    from datetime import datetime, timedelta
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_count = MaterialSKU.query.filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.created_at >= today_start
    ).count()

    # 在售物料（active 状态）
    active_count = dict(status_stats).get('active', 0)

    return jsonify({
        'code': 200,
        'data': {
            'total': total,
            'active': active_count,
            'warning': low_stock,
            'today': today_count
        }
    })


# ========== 品牌列表 ==========

@material_sku_bp.route('/brands', methods=['GET'])
@jwt_required_v2
def get_brands(current_user):
    """获取品牌列表（从物料中提取）"""
    from sqlalchemy import distinct
    brands = db.session.query(distinct(MaterialSKU.brand)).filter(
        MaterialSKU.brand.isnot(None),
        MaterialSKU.brand != '',
        MaterialSKU.is_deleted == False
    ).order_by(MaterialSKU.brand).all()
    
    return jsonify({
        'code': 200,
        'data': [b[0] for b in brands if b[0]]
    })


@material_sku_bp.route('/filter-options', methods=['GET'])
def get_filter_options():
    """获取筛选栏所有下拉选项（从数据库真实数据动态提取）"""
    from sqlalchemy import distinct

    # 品牌
    brands_raw = db.session.query(distinct(MaterialSKU.brand)).filter(
        MaterialSKU.brand.isnot(None),
        MaterialSKU.brand != '',
        MaterialSKU.is_deleted == False
    ).order_by(MaterialSKU.brand).all()
    brands = [b[0] for b in brands_raw if b[0]]
    
    # 单位
    units_raw = db.session.query(distinct(MaterialSKU.unit)).filter(
        MaterialSKU.unit.isnot(None),
        MaterialSKU.unit != '',
        MaterialSKU.is_deleted == False
    ).order_by(MaterialSKU.unit).all()
    units = [u[0] for u in units_raw if u[0]]
    
    # 供应链
    sc_raw = db.session.query(distinct(MaterialSKU.supply_chain)).filter(
        MaterialSKU.supply_chain.isnot(None),
        MaterialSKU.supply_chain != '',
        MaterialSKU.is_deleted == False
    ).order_by(MaterialSKU.supply_chain).all()
    supply_chains = [s[0] for s in sc_raw if s[0]]
    
    # 环保等级
    env_raw = db.session.query(distinct(MaterialSKU.env_level)).filter(
        MaterialSKU.env_level.isnot(None),
        MaterialSKU.env_level != '',
        MaterialSKU.is_deleted == False
    ).order_by(MaterialSKU.env_level).all()
    env_levels = [e[0] for e in env_raw if e[0]]
    
    # 状态（带中文标签）
    status_raw = db.session.query(distinct(MaterialSKU.status)).filter(
        MaterialSKU.status.isnot(None),
        MaterialSKU.is_deleted == False
    ).order_by(MaterialSKU.status).all()
    status_map = {
        'active': '在售',
        'draft': '草稿',
        'discontinued': '停售'
    }
    statuses = [{'value': s[0], 'label': status_map.get(s[0], s[0])} for s in status_raw if s[0]]
    
    return jsonify({
        'code': 200,
        'data': {
            'brands': brands,
            'units': units,
            'supply_chains': supply_chains,
            'env_levels': env_levels,
            'statuses': statuses
        }
    })


# ========== 辅助函数 ==========

def generate_sku_code():
    """生成SKU编码"""
    import random
    import string
    prefix = 'SKU'
    suffix = ''.join(random.choices(string.digits, k=6))
    return f"{prefix}{suffix}"


# ========== Public Display Config ==========

@material_sku_bp.route('/public-config', methods=['GET'])
@jwt_required_v2
def get_public_config(current_user):
    """Get public display config for frontend"""
    # Get categories with public materials
    public_categories = db.session.query(
        MaterialSKU.category_id
    ).filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.is_public == True,
        MaterialSKU.category_id.isnot(None)
    ).distinct().all()
    
    # Get brands with public materials
    public_brands = db.session.query(
        MaterialSKU.brand
    ).filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.is_public == True,
        MaterialSKU.brand.isnot(None),
        MaterialSKU.brand != ''
    ).distinct().all()
    
    # Count total public materials
    total_public = MaterialSKU.query.filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.is_public == True
    ).count()
    
    return jsonify({
        'code': 200,
        'data': {
            'categories': [c[0] for c in public_categories if c[0]],
            'brands': [b[0] for b in public_brands if b[0]],
            'total': total_public
        }
    })


@material_sku_bp.route('/public-config', methods=['POST'])
@jwt_required_v2
def save_public_config(current_user):
    """Save public display config - batch update is_public field"""
    from sqlalchemy import or_
    
    data = request.get_json() or {}
    categories = data.get('categories', [])
    brands = data.get('brands', [])
    
    # First, set all materials to not public
    MaterialSKU.query.update({'is_public': False})
    
    # Then, set selected categories and brands to public
    if categories or brands:
        query = MaterialSKU.query.filter(MaterialSKU.is_deleted == False)
        
        conditions = []
        if categories:
            conditions.append(MaterialSKU.category_id.in_(categories))
        if brands:
            conditions.append(MaterialSKU.brand.in_(brands))
        
        if conditions:
            query.filter(or_(*conditions)).update({'is_public': True}, synchronize_session=False)
    
    db.session.commit()
    
    # Return updated stats
    total_public = MaterialSKU.query.filter(
        MaterialSKU.is_deleted == False,
        MaterialSKU.is_public == True
    ).count()
    
    return jsonify({
        'code': 200,
        'message': 'Saved successfully',
        'data': {
            'total': total_public
        }
    })


# ========== 门店私库物料 ==========

@material_sku_bp.route('/private', methods=['GET'])
@jwt_required_v2
def get_private_materials(current_user):
    from app.models.material_store_private import MaterialStorePrivate
    
    store_id = current_user.get('store_id') or 1
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('pageSize', 50, type=int)
    keyword = request.args.get('keyword', '')
    
    query = MaterialStorePrivate.query.filter(MaterialStorePrivate.store_id == store_id)
    
    if keyword:
        query = query.filter(db.or_(
            MaterialStorePrivate.name.ilike(f'%{keyword}%'),
            MaterialStorePrivate.sku_code.ilike(f'%{keyword}%')
        ))
    
    pagination = query.order_by(MaterialStorePrivate.updated_at.desc()).paginate(page, page_size, False)
    
    return jsonify({
        'code': 200,
        'data': [m.to_dict() for m in pagination.items],
        'total': pagination.total
    })


@material_sku_bp.route('/private', methods=['POST'])
@jwt_required_v2
def create_private_material(current_user):
    from app.models.material_store_private import MaterialStorePrivate
    
    data = request.get_json()
    store_id = current_user.get('store_id') or 1
    
    m = MaterialStorePrivate(
        store_id=store_id,
        name=data['name'],
        sku_code=data.get('sku_code'),
        category_id=data.get('category_id'),
        spec=data.get('spec', ''),
        material_type=data.get('material_type', ''),
        brand=data.get('brand', ''),
        color=data.get('color', ''),
        unit=data.get('unit', ''),
        unit_price=data.get('unit_price', 0),
        remark=data.get('remark', ''),
        created_by=current_user.id
    )
    
    db.session.add(m)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': 'created', 'data': m.to_dict()})


@material_sku_bp.route('/private/<int:id>', methods=['PUT'])
@jwt_required_v2
def update_private_material(current_user, id):
    from app.models.material_store_private import MaterialStorePrivate
    
    m = MaterialStorePrivate.query.get_or_404(id)
    data = request.get_json()
    
    for field in ['name', 'sku_code', 'category_id', 'spec', 'material_type', 'brand', 'color', 'unit', 'unit_price', 'remark']:
        if field in data:
            setattr(m, field, data[field])
    
    db.session.commit()
    
    return jsonify({'code': 200, 'message': 'updated', 'data': m.to_dict()})


@material_sku_bp.route('/private/<int:id>', methods=['DELETE'])
@jwt_required_v2
def delete_private_material(current_user, id):
    from app.models.material_store_private import MaterialStorePrivate
    
    m = MaterialStorePrivate.query.get_or_404(id)
    db.session.delete(m)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': 'deleted'})


# ========== Excel 批量导入物料 ==========

@material_sku_bp.route('/import-template', methods=['GET'])
@jwt_required_v2
def download_material_import_template(current_user):
    """下载物料导入 Excel 模板"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '物料导入模板'
    
    # 表头样式
    header_fill = PatternFill(start_color='8B5A2B', end_color='8B5A2B', fill_type='solid')
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    cell_font = Font(name='微软雅黑', size=10)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    headers = ['SKU编码*', '名称*', '分类', '品牌', '型号', '规格', '材质', '产地', '成本价', '销售价', '单位', '计价方式', '库存', '状态', '是否公开', '备注']
    col_widths = [15, 20, 12, 12, 12, 15, 10, 10, 10, 10, 8, 10, 8, 8, 8, 20]
    
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    
    ws.row_dimensions[1].height = 28
    
    # 示例行
    example_data = ['D&B-KT-001', '定制衣柜柜体', '固装家具', 'D&B帝标', 'WGB-120', '宽1200*深600*高2400', '实木颗粒板', '成都', '800', '1200', '㎡', '展开报价', '100', '在售', '是', 'E0级环保板材']
    for col, val in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = cell_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 字段说明sheet
    ws2 = wb.create_sheet('字段说明')
    ws2['A1'] = '字段'
    ws2['B1'] = '说明/示例'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    notes = [
        ('SKU编码', '必填，唯一标识'),
        ('名称', '必填，物料名称'),
        ('分类', '物料分类名称'),
        ('计价方式', '展开报价/投影报价/延米报价/一口价'),
        ('状态', '在售/停售/下架'),
        ('是否公开', '是/否'),
    ]
    for i, (k, v) in enumerate(notes, 2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 30
    
    ws.freeze_panes = 'A2'
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='物料导入模板.xlsx'
    )


@material_sku_bp.route('/import-excel', methods=['POST'])
@jwt_required_v2
def import_materials_from_excel(current_user):
    """Excel 批量导入物料
    
    Excel 列头支持（自动识别）：
    SKU编码/编码, 名称/物料名称, 分类/物料分类, 品牌, 型号, 规格, 材质, 产地,
    成本价, 销售价/单价, 单位, 计价方式, 库存, 状态, 是否公开, 备注
    """
    import io
    import openpyxl
    from sqlalchemy import or_
    
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传 Excel 文件'})
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 400, 'message': '文件名为空'})
    
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {'.xlsx', '.xls'}:
        return jsonify({'code': 400, 'message': f'不支持的文件格式 {ext}，请上传 xlsx/xls'})
    
    try:
        # 读取 Excel
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows or len(rows) < 2:
            return jsonify({'code': 400, 'message': 'Excel 文件为空或无数据'})
        
        # 解析表头
        headers = [str(c).strip() if c else '' for c in rows[0]]
        
        # 列名映射
        COLUMN_MAP = {
            'SKU编码': 'sku_code', '编码': 'sku_code', '物料编码': 'sku_code',
            '名称': 'name', '物料名称': 'name', '产品名称': 'name',
            '分类': 'category_name', '物料分类': 'category_name', '分类名称': 'category_name',
            '品牌': 'brand',
            '型号': 'model', '产品型号': 'model',
            '规格': 'specification', '产品规格': 'specification',
            '材质': 'material', '材料': 'material',
            '产地': 'origin', '生产地': 'origin',
            '成本价': 'cost_price', '进货价': 'cost_price',
            '销售价': 'unit_price', '单价': 'unit_price', '零售价': 'unit_price',
            '单位': 'unit', '计量单位': 'unit',
            '计价方式': 'calc_type', '计价类型': 'calc_type',
            '库存': 'stock', '库存数量': 'stock',
            '状态': 'status', '物料状态': 'status',
            '是否公开': 'is_public', '公开': 'is_public',
            '备注': 'remark', '说明': 'remark',
        }
        
        col_map = {}
        for i, h in enumerate(headers):
            h = str(h).strip()
            if h in COLUMN_MAP:
                col_map[i] = COLUMN_MAP[h]
        
        if 'sku_code' not in col_map.values() and 'name' not in col_map.values():
            return jsonify({'code': 400, 'message': 'Excel 表头无法识别，请使用标准模板列名'})
        
        imported = 0
        updated = 0
        skipped = 0
        errors = []
        
        for idx, row in enumerate(rows[1:], 2):  # 从第2行开始
            try:
                row_data = {}
                for i in col_map:
                    val = row[i] if i < len(row) else None
                    row_data[col_map[i]] = str(val).strip() if val is not None else ''
                
                sku_code = row_data.get('sku_code', '').strip()
                name = row_data.get('name', '').strip()
                
                if not name:
                    skipped += 1
                    continue
                
                # 查找分类ID
                category_id = None
                category_name = row_data.get('category_name', '').strip()
                if category_name:
                    cat = MaterialCategory.query.filter(
                        MaterialCategory.name == category_name,
                        or_(MaterialCategory.is_deleted == False, MaterialCategory.is_deleted.is_(None))
                    ).first()
                    if cat:
                        category_id = cat.id
                
                # 处理数值字段
                def parse_float(val):
                    try:
                        return float(val) if val and val.replace('.', '').replace('-', '').isdigit() else 0
                    except:
                        return 0
                
                cost_price = parse_float(row_data.get('cost_price', '0'))
                unit_price = parse_float(row_data.get('unit_price', '0'))
                stock = parse_float(row_data.get('stock', '0'))
                
                # 处理计价方式
                calc_type = row_data.get('calc_type', '一口价').strip()
                if calc_type not in CALC_TYPES.values():
                    # 尝试匹配
                    calc_type = '一口价'
                    for k, v in CALC_TYPES.items():
                        if v == row_data.get('calc_type', ''):
                            calc_type = v
                            break
                
                # 处理状态
                status = row_data.get('status', '在售').strip()
                if status not in SKU_STATUS.values():
                    status = '在售'
                
                # 处理是否公开
                is_public = row_data.get('is_public', '否').strip() in ('是', 'YES', 'Y', 'TRUE', '1', '公开')
                
                # 检查是否已存在（按SKU编码或名称）
                existing = None
                if sku_code:
                    existing = MaterialSKU.query.filter_by(sku_code=sku_code).first()
                if not existing and name:
                    existing = MaterialSKU.query.filter_by(name=name).first()
                
                if existing:
                    # 更新
                    existing.name = name
                    existing.category_id = category_id
                    existing.brand = row_data.get('brand') or existing.brand
                    existing.model = row_data.get('model') or existing.model
                    existing.specification = row_data.get('specification') or existing.specification
                    existing.material = row_data.get('material') or existing.material
                    existing.origin = row_data.get('origin') or existing.origin
                    existing.cost_price = cost_price or existing.cost_price
                    existing.unit_price = unit_price or existing.unit_price
                    existing.unit = row_data.get('unit') or existing.unit
                    existing.calc_type = calc_type
                    existing.stock = stock
                    existing.status = status
                    existing.is_public = is_public
                    existing.remark = row_data.get('remark') or existing.remark
                    existing.updated_at = datetime.utcnow()
                    updated += 1
                else:
                    # 新建
                    m = MaterialSKU(
                        sku_code=sku_code or f'SKU{datetime.now().strftime("%Y%m%d%H%M%S")}{imported+1:03d}',
                        name=name,
                        category_id=category_id,
                        brand=row_data.get('brand'),
                        model=row_data.get('model'),
                        specification=row_data.get('specification'),
                        material=row_data.get('material'),
                        origin=row_data.get('origin'),
                        cost_price=cost_price,
                        unit_price=unit_price,
                        unit=row_data.get('unit'),
                        calc_type=calc_type,
                        stock=stock,
                        status=status,
                        is_public=is_public,
                        remark=row_data.get('remark'),
                        created_by=current_user.id
                    )
                    db.session.add(m)
                    imported += 1
                
            except Exception as e:
                errors.append(f'第{idx}行：{str(e)}')
                skipped += 1
        
        db.session.commit()
        
        msg = f'导入完成：新增 {imported} 条，更新 {updated} 条'
        if skipped > 0:
            msg += f'，跳过 {skipped} 条'
        if errors:
            msg += f'，错误：' + '；'.join(errors[:5])
            if len(errors) > 5:
                msg += f'... 共{len(errors)}条错误'
        
        return jsonify({'code': 200, 'message': msg, 'data': {'imported': imported, 'updated': updated, 'skipped': skipped}})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'})

@material_sku_bp.route('/batch-import', methods=['POST'])
def batch_import_materials():
    """批量导入物料，自动创建不重复的L2分类"""
    data = request.get_json()
    if not data or 'items' not in data:
        return jsonify({'code': 400, 'message': '请提供物料列表(items)'})

    items = data['items']
    if not items or not isinstance(items, list):
        return jsonify({'code': 400, 'message': '物料列表不能为空'})

    # 加载所有现有分类
    all_cats = MaterialCategory.query.filter_by(is_deleted=False).all()
    cat_name_map = {}
    for c in all_cats:
        key = f"{c.parent_id or 0}|{c.name}"
        cat_name_map[key] = c
    l1_name_map = {c.name: c for c in all_cats if c.level == 1 and not c.is_deleted}

    results = {'created': 0, 'skipped': 0, 'errors': [], 'new_categories': []}

    for idx, item in enumerate(items):
        try:
            l1_name = item.get('category_level1', '').strip()
            l2_name = item.get('category_level2', '').strip()

            if not l1_name:
                results['errors'].append(f'第{idx+1}行: 缺少一级分类')
                continue

            # 查找或创建 L1
            l1_cat = l1_name_map.get(l1_name)
            if not l1_cat:
                l1_cat = MaterialCategory(
                    name=l1_name, level=1, sort_order=len(l1_name_map) + 1
                )
                db.session.add(l1_cat)
                db.session.flush()
                l1_name_map[l1_name] = l1_cat
                cat_name_map[f"0|{l1_name}"] = l1_cat
                results['new_categories'].append({'level': 1, 'name': l1_name, 'id': l1_cat.id})

            # 查找或创建 L2
            category_id = l1_cat.id
            if l2_name:
                key = f"{l1_cat.id}|{l2_name}"
                l2_cat = cat_name_map.get(key)
                if not l2_cat:
                    max_sort = db.session.query(db.func.max(MaterialCategory.sort_order)).filter_by(
                        parent_id=l1_cat.id, is_deleted=False
                    ).scalar() or 0
                    l2_cat = MaterialCategory(
                        name=l2_name, parent_id=l1_cat.id, level=2, sort_order=max_sort + 1
                    )
                    db.session.add(l2_cat)
                    db.session.flush()
                    cat_name_map[key] = l2_cat
                    results['new_categories'].append({'level': 2, 'name': l2_name, 'parent': l1_name, 'id': l2_cat.id})
                category_id = l2_cat.id

            # SKU编码去重
            sku_code = item.get('sku_code', '').strip()
            if sku_code:
                existing = MaterialSKU.query.filter_by(sku_code=sku_code).first()
                if existing:
                    results['skipped'] += 1
                    continue

            # 创建SKU
            sku = MaterialSKU(
                sku_code=sku_code or f"AUTO-{idx+1:05d}",
                name=item.get('name', '').strip(),
                category_id=category_id,
                brand=item.get('brand', ''),
                model=item.get('model', ''),
                specification=item.get('specification', ''),
                material=item.get('material', ''),
                origin=item.get('origin', ''),
                main_image=item.get('main_image', ''),
                color_name=item.get('color_name', ''),
                env_level=item.get('env_level', '合格'),
                supply_chain=item.get('supply_chain', '直供'),
                cost_price=float(item.get('cost_price', 0) or 0),
                sale_price=float(item.get('sale_price', 0) or 0),
                market_price=float(item.get('market_price', 0) or 0) if item.get('market_price') else None,
                unit=item.get('unit', '件'),
                calc_type=item.get('calc_type', 'quantity'),
                stock_quantity=int(item.get('stock_quantity', 0) or 0),
                description=item.get('description', ''),
                status=item.get('status', 'active'),
                is_public=item.get('is_public', True),
            )
            db.session.add(sku)
            results['created'] += 1

        except Exception as e:
            results['errors'].append(f'第{idx+1}行: {str(e)}')
            db.session.rollback()
            continue

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'批量导入失败: {str(e)}'})

    return jsonify({
        'code': 200,
        'message': f'批量导入完成: 创建{results["created"]}条, 跳过{results["skipped"]}条, 新分类{len(results["new_categories"])}个',
        'data': results
    })
